
import random
from collections import defaultdict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd

import openpyxl
class CSPFinalTakvimi:
    def __init__(self, ders_kisitlari, gunler, saatler, siniflar, ders_bolum_sayilari,bolumler):
        self.ders_kisitlari = ders_kisitlari
        self.gunler = gunler
        self.saatler = saatler
        self.siniflar = siniflar
        self.ders_bolum_sayilari = ders_bolum_sayilari
        self.bolumler=bolumler
        self.domains = {
            ders: sorted(
            [
                (g, s, tuple(sinif_grubu))
                for g in gunler
                for s in saatler
                for sinif_grubu in self.uygun_siniflari_bul(ders)
            ],
            key=lambda x: random.random()
            )
            for ders in ders_kisitlari.index
        }

    def uygun_siniflari_bul(self, ders):
        ders_bilgisi = self.ders_bolum_sayilari[ders]
        gerekli_kapasite = ders_bilgisi["kapasite"]
        
        sınıflar = self.siniflar[:]
        random.shuffle(sınıflar) 

        toplam_kapasite = 0
        atanan_sınıflar = []

        for sınıf_adı, sınıf_kapasite in sınıflar:
            if toplam_kapasite < gerekli_kapasite:
                atanan_sınıflar.append((sınıf_adı, sınıf_kapasite))
                toplam_kapasite += sınıf_kapasite
            else:
                break

        if toplam_kapasite >= gerekli_kapasite:
            return atanan_sınıflar
        else:
            return None

    def uygun_mu(self, atamalar):
        zaman_sinif_doluluk = set()
        ders_zaman_map = {}
        
        #aynı gün ve saatte aynı sınıfa birden fazla sınıf atanamaz
        for ders, (gun, saat, sinif_grubu) in atamalar.items():
            for sinif in sinif_grubu:
                if (gun, saat, sinif) in zaman_sinif_doluluk:
                    return False
                zaman_sinif_doluluk.add((gun, saat, sinif))
        #her ders sadece bir zamanda olmalıdır
            if ders in ders_zaman_map:
                onceki_gun, onceki_saat = ders_zaman_map[ders]
                if gun != onceki_gun or saat != onceki_saat:
                    return False
            else:
                ders_zaman_map[ders] = (gun, saat)
        
        #ders ilişkisi -1000 dersleri aynı güne konmuş ise kabul edilmez
        dersler = list(atamalar.keys())
        for i in range(len(dersler)):
            for j in range(i + 1, len(dersler)):
                d1, d2 = dersler[i], dersler[j]
                g1 = atamalar[d1][0]
                g2 = atamalar[d2][0]
                if d1 != d2 and g1 == g2 and self.ders_kisitlari.loc[d1, d2] == -1000:
                    return False
        return True

    #daha iyi çözüm olup olmadığı kontrol edilir
    def puanla(self, atamalar):
        puan = 1000
        dersler = list(atamalar.keys())

        #derslerin arasındaki ilişkiye göre 
        for i in range(len(dersler)):
            for j in range(i + 1, len(dersler)):
                d1, d2 = dersler[i], dersler[j]
                g1 = atamalar[d1][0]
                g2 = atamalar[d2][0]
                if d1 == d2:
                    continue
                k = self.ders_kisitlari.loc[d1, d2]
                if g1 == g2:
                    if k == -500:
                        puan -= 100
                    elif k == -200:
                        puan -= 50

        gun_ders_map = defaultdict(set)
        for ders, (gun, _, _) in atamalar.items():
            gun = int(gun)
            gun_ders_map[gun].add(ders)
            
        #derslerin üç gün üst üste olup olmaması 
        gunler = sorted([int(g) for g in gun_ders_map.keys()])
        for i in range(len(gunler) - 2):
            g1, g2, g3 = gunler[i], gunler[i + 1], gunler[i + 2]
            if g2 - g1 == 1 and g3 - g2 == 1:
                for d1 in gun_ders_map[g1]:
                    for d2 in gun_ders_map[g2]:
                        for d3 in gun_ders_map[g3]:
                            if (
                                self.ders_kisitlari.loc[d1, d2] == -1000 and
                                self.ders_kisitlari.loc[d2, d3] == -1000 and
                                self.ders_kisitlari.loc[d1, d3] == -1000
                            ):
                                puan -= 300
        
        gun_sirasi = {gun: i for i, gun in enumerate(self.gunler)}
        saat_sirasi = {saat: i for i, saat in enumerate(self.saatler)}

        #aynı bölümden olan derslerin aynı gün içinde yakın olması
        for i in range(len(dersler)):
            for j in range(i + 1, len(dersler)):
                d1, d2 = dersler[i], dersler[j]
                g1, s1 = atamalar[d1][0], atamalar[d1][1]
                g2, s2 = atamalar[d2][0], atamalar[d2][1]
               
                if self.ayni_bolumden(d1, d2):
                    g1_index = gun_sirasi.get(g1, 0)
                    g2_index = gun_sirasi.get(g2, 0)
                    s1_index = saat_sirasi.get(s1, 0)
                    s2_index = saat_sirasi.get(s2, 0)
                    
                    gun_farki = abs(g1_index - g2_index)
                    saat_farki = abs(s1_index - s2_index)
                    if gun_farki == 0 and saat_farki <= 1:
                        puan -= 100
        return puan
    def ayni_bolumden(self, ders1, ders2):
        for bolum, donemler in self.bolumler.items():
            tum_dersler = [d for dersler in donemler.values() for d in dersler]
            if ders1 in tum_dersler and ders2 in tum_dersler:
                return True
        return False

    def backtrack(self, atamalar, kalan_dersler, max_depth=1000, current_depth=0):
        if not kalan_dersler:
           return atamalar

        if current_depth > max_depth:
            return None

        ders = kalan_dersler[0]
        for (gun, saat, sinif_grubu) in self.domains[ders]:
            yeni_atamalar = atamalar.copy()
            yeni_atamalar[ders] = (gun, saat, sinif_grubu)
            if self.uygun_mu(yeni_atamalar):
                sonuc = self.backtrack(yeni_atamalar, kalan_dersler[1:], max_depth, current_depth + 1)
                if sonuc:
                    return sonuc
                return None


    def solve(self):
        dersler = list(self.domains.keys())
        en_iyi = None
        en_yuksek_puan = float('-inf')

        for _ in range(100):
            random.shuffle(dersler)
            cozum = self.backtrack({}, dersler)
            if cozum:
                puan = self.puanla(cozum)
                if puan > en_yuksek_puan:
                    en_yuksek_puan = puan
                    en_iyi = cozum

        return en_iyi


class FinalTakvimUygulamasi:
    def __init__(self, root):
        self.root = root
        self.root.title("Final Takvimi Planlama")
        self.bolumler = {}  
        self.cap_bolumler = []  
        self.siniflar = []  

        self.main_frame = tk.Frame(root)
        self.main_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
     
        self.sol_frame = tk.Frame(self.main_frame)
        self.sol_frame.pack(side=tk.LEFT, padx=10, fill=tk.Y)
     
        self.sag_frame = tk.Frame(self.main_frame)
        self.sag_frame.pack(side=tk.RIGHT, padx=10, fill=tk.Y)

        self.excel_frame = tk.Frame(self.sol_frame)
        self.excel_frame.pack(pady=10, fill=tk.X)
        tk.Button(self.excel_frame, text="Excel'den Veri Yükle", 
                 command=self.excelden_veri_yukle).pack(fill=tk.X)

        self.bolum_frame = tk.Frame(self.sol_frame)
        self.bolum_frame.pack(pady=10, fill=tk.X)
        tk.Label(self.bolum_frame, text="Bölüm Adı:").pack(side=tk.LEFT)
        self.bolum_entry = tk.Entry(self.bolum_frame)
        self.bolum_entry.pack(side=tk.LEFT)
        tk.Button(self.bolum_frame, text="Ekle", command=self.bolum_ekle).pack(side=tk.LEFT)
        tk.Button(self.bolum_frame, text="Sil", command=self.bolum_sil).pack(side=tk.LEFT)

        self.bolum_listbox = tk.Listbox(self.sol_frame, height=5)
        self.bolum_listbox.pack(pady=10, fill=tk.X)
        self.bolum_listbox.bind("<<ListboxSelect>>", self.bolum_secildi)
 
        self.donem_frame = tk.Frame(self.sol_frame)
        self.donem_frame.pack(pady=10, fill=tk.X)
        tk.Label(self.donem_frame, text="Dönem Seç:").pack(side=tk.LEFT)
        self.donem_var = tk.StringVar()
        self.donem_menu = ttk.Combobox(self.donem_frame, textvariable=self.donem_var)
 
        donem_secenekleri = ["Tekrar/Seçmeli"]  
        donem_secenekleri.extend([f"{i}. Yarıyıl" for i in range(1, 12)])  
        self.donem_menu["values"] = donem_secenekleri
        
        self.donem_menu.pack(side=tk.LEFT)
        tk.Button(self.donem_frame, text="Dönem Ekle", command=self.donem_ekle).pack(side=tk.LEFT)
        tk.Button(self.donem_frame, text="Dönem Sil", command=self.donem_sil).pack(side=tk.LEFT)
   
        self.donem_listbox = tk.Listbox(self.sol_frame, height=5)
        self.donem_listbox.pack(pady=10, fill=tk.X)
        self.donem_listbox.bind("<<ListboxSelect>>", self.donem_secildi_listbox)
        
        self.ders_frame = tk.Frame(self.sol_frame)
        tk.Label(self.ders_frame, text="Ders Adı:").pack(side=tk.LEFT)
        self.ders_entry = tk.Entry(self.ders_frame)
        self.ders_entry.pack(side=tk.LEFT)
        tk.Button(self.ders_frame, text="Ders Ekle", command=self.ders_ekle).pack(side=tk.LEFT)
        tk.Button(self.ders_frame, text="Sil", command=self.ders_sil).pack(side=tk.LEFT)
        self.ders_frame.pack_forget() 
        
        self.ders_listbox = tk.Listbox(self.sol_frame, height=5)
        self.ders_listbox.pack(pady=5, fill=tk.X)
 
        self.cap_frame = tk.Frame(self.sag_frame)
        self.cap_frame.pack(pady=10, fill=tk.X)
        tk.Label(self.cap_frame, text="Birinci Çap Bölümü:").pack()
        self.cap_bolum_var1 = tk.StringVar()
        self.cap_bolum_menu1 = ttk.Combobox(self.cap_frame, textvariable=self.cap_bolum_var1)
        self.cap_bolum_menu1.pack()

        tk.Label(self.cap_frame, text="İkinci Çap Bölümü:").pack()
        self.cap_bolum_var2 = tk.StringVar()
        self.cap_bolum_menu2 = ttk.Combobox(self.cap_frame, textvariable=self.cap_bolum_var2)
        self.cap_bolum_menu2.pack()
        
        tk.Button(self.cap_frame, text="Ekle", command=self.cap_ders_ekle).pack()
        tk.Button(self.cap_frame, text="Sil", command=self.cap_ders_sil).pack()
        
        self.cap_listbox = tk.Listbox(self.sag_frame, height=5)
        self.cap_listbox.pack(pady=5, fill=tk.X)
        
        # Sınıf ekleme alanı
        self.sinif_frame = tk.Frame(self.sag_frame)
        self.sinif_frame.pack(pady=10, fill=tk.X)
        tk.Label(self.sinif_frame, text="Sınıf Adı:").pack(side=tk.LEFT)
        self.sinif_entry = tk.Entry(self.sinif_frame)
        self.sinif_entry.pack(side=tk.LEFT)
        tk.Label(self.sinif_frame, text="Kapasite:").pack(side=tk.LEFT)
        self.kapasite_entry = tk.Entry(self.sinif_frame, width=5)
        self.kapasite_entry.pack(side=tk.LEFT)
        tk.Button(self.sinif_frame, text="Ekle", command=self.sinif_ekle).pack(side=tk.LEFT)
        tk.Button(self.sinif_frame, text="Sil", command=self.sinif_sil).pack(side=tk.LEFT)
        
        self.sinif_listbox = tk.Listbox(self.sag_frame, height=5)
        self.sinif_listbox.pack(pady=5, fill=tk.X)
  
        self.takvim_buton = tk.Button(self.sag_frame, text="Final Takvimi Oluştur", command=self.takvim_hazirla)
        self.takvim_buton.pack(pady=10)
        
    def excelden_veri_yukle(self):
        dosya_yolu = filedialog.askopenfilename(
            title="Excel Dosyası Seçin",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )
        
        if not dosya_yolu:
            return
            
        try:

            wb = openpyxl.load_workbook(dosya_yolu)
            
            self.bolumler = {}
            self.bolum_listbox.delete(0, tk.END)
            self.siniflar = []
            self.sinif_listbox.delete(0, tk.END)
         
            if "sınıflar" in wb.sheetnames:
                sheet = wb["sınıflar"]
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    if row[0] and row[1]:
                        sinif_adi = str(row[0])
                        kapasite = int(row[1]) if str(row[1]).isdigit() else 0
                        if kapasite > 0:
                            self.siniflar.append((sinif_adi, kapasite))
                            self.sinif_listbox.insert(tk.END, f"{sinif_adi} - {kapasite}")
   
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() != "sınıflar":  
                    bolum_adi = sheet_name
                    sheet = wb[sheet_name]
                
         
                    self.bolumler[bolum_adi] = {}
                    self.bolum_listbox.insert(tk.END, bolum_adi)
                
     
                    donemler = []
                    for col in sheet.iter_cols(min_row=1, max_row=1):
                        donem = col[0].value
                        if donem:  
                            donemler.append(donem)

                    for i, donem in enumerate(donemler):
                        dersler = []
                        for cell in sheet.iter_rows(min_row=2, min_col=i+1, max_col=i+1):
                            if cell[0].value: 
                                dersler.append(str(cell[0].value))
                    
                        if dersler:
                            self.bolumler[bolum_adi][donem] = dersler
 
            self.cap_bolum_menu1["values"] = list(self.bolumler.keys())
            self.cap_bolum_menu2["values"] = list(self.bolumler.keys())
            
            messagebox.showinfo("Başarılı", "Excel dosyası başarıyla yüklendi!")
            
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası okunurken hata oluştu:\n{str(e)}")
    

    def bolum_ekle(self):
        bolum_adi = self.bolum_entry.get()
        if bolum_adi and bolum_adi not in self.bolumler:
            self.bolumler[bolum_adi] = {}  
            self.bolum_listbox.insert(tk.END, bolum_adi)
            self.cap_bolum_menu1["values"] = list(self.bolumler.keys())
            self.cap_bolum_menu2["values"] = list(self.bolumler.keys())
        self.bolum_entry.delete(0, tk.END)
    
    def bolum_sil(self):
        selection = self.bolum_listbox.curselection()
        if selection:
            bolum_adi = self.bolum_listbox.get(selection[0])
            del self.bolumler[bolum_adi]  
            self.bolum_listbox.delete(selection[0])
            self.cap_bolum_menu1["values"] = list(self.bolumler.keys())
            self.cap_bolum_menu2["values"] = list(self.bolumler.keys())
            self.donem_listbox.delete(0, tk.END)
            self.ders_listbox.delete(0, tk.END)
    
    def bolum_secildi(self, event):
        selection = self.bolum_listbox.curselection()
        if selection:
            self.selected_bolum = self.bolum_listbox.get(selection[0])

            self.donem_listbox.delete(0, tk.END)
            if self.selected_bolum in self.bolumler:
                for donem in self.bolumler[self.selected_bolum]:
                    self.donem_listbox.insert(tk.END, donem)
            self.ders_frame.pack_forget()  
    
    def donem_ekle(self):
        donem = self.donem_var.get()
        if self.selected_bolum and donem:
            if donem not in self.bolumler[self.selected_bolum]:
                self.bolumler[self.selected_bolum][donem] = [] 
                self.donem_listbox.insert(tk.END, donem) 
            self.donem_var.set('')  
    
    def donem_sil(self):
        selection = self.donem_listbox.curselection()
        if selection and self.selected_bolum:
            donem = self.donem_listbox.get(selection[0])
            del self.bolumler[self.selected_bolum][donem]  
            self.donem_listbox.delete(selection[0])  
            self.ders_listbox.delete(0, tk.END)  
            self.ders_frame.pack_forget()  
    
    def donem_secildi_listbox(self, event):
        selection = self.donem_listbox.curselection()
        if selection:
            self.selected_donem = self.donem_listbox.get(selection[0])
            self.ders_frame.pack(fill=tk.X)
            self.ders_listbox.delete(0, tk.END)
            if self.selected_donem in self.bolumler[self.selected_bolum]:
                for ders in self.bolumler[self.selected_bolum][self.selected_donem]:
                    self.ders_listbox.insert(tk.END, ders)
    
    def ders_ekle(self):
        ders_adi = self.ders_entry.get()
        if self.selected_bolum and self.selected_donem and ders_adi:
            if self.selected_donem not in self.bolumler[self.selected_bolum]:
                self.bolumler[self.selected_bolum][self.selected_donem] = []
            self.bolumler[self.selected_bolum][self.selected_donem].append(ders_adi)
            self.ders_listbox.insert(tk.END, ders_adi)
            self.ders_entry.delete(0, tk.END)
    
    def ders_sil(self):
        selection = self.ders_listbox.curselection()
        if selection:
            ders_adi = self.ders_listbox.get(selection[0])
            self.bolumler[self.selected_bolum][self.selected_donem].remove(ders_adi)
            self.ders_listbox.delete(selection[0])
    
    def cap_ders_ekle(self):
        bolum1 = self.cap_bolum_var1.get()
        bolum2 = self.cap_bolum_var2.get()
        if bolum1 in self.bolumler and bolum2 in self.bolumler:
            self.cap_listbox.insert(tk.END, f"{bolum1} - {bolum2}")
            self.cap_bolumler.append((bolum1, bolum2)) 
    
    def cap_ders_sil(self):
        selection = self.cap_listbox.curselection()
        if selection:
            cap_birim = self.cap_listbox.get(selection[0])
            self.cap_listbox.delete(selection[0])
            bolum1, bolum2 = cap_birim.split(" - ")
            self.cap_bolumler.remove((bolum1, bolum2))
    
    def sinif_ekle(self):
        sinif_adi = self.sinif_entry.get()
        kapasite = self.kapasite_entry.get()
        if sinif_adi and kapasite.isdigit():
            self.siniflar.append((sinif_adi, int(kapasite)))
            self.sinif_listbox.insert(tk.END, f"{sinif_adi} - {kapasite}")
            self.sinif_entry.delete(0, tk.END)
            self.kapasite_entry.delete(0, tk.END)
    
    def sinif_sil(self):
        selection = self.sinif_listbox.curselection()
        if selection:
            self.siniflar.pop(selection[0])
            self.sinif_listbox.delete(selection[0])

    def ders_iliski_olustur(self, ceza_degeri=-200, ek_ders_ceza=-400):
        dersler = list(set(
            ders for bolum in self.bolumler.values()
            for donem_dersleri in bolum.values()
            for ders in donem_dersleri
        ))
        

        ders_kisitlari = pd.DataFrame(0, index=dersler, columns=dersler)

        for bolum, donemler in self.bolumler.items():
            tekrar_secmeli_dersler = donemler.get("Tekrar/Seçmeli", [])

            diger_dersler = [
                ders for donem, donem_dersleri in donemler.items()
                if donem != "Tekrar/Seçmeli"
                for ders in donem_dersleri
            ]

            for tekrar_ders in tekrar_secmeli_dersler:
                for diger_ders in diger_dersler:
                    if tekrar_ders != diger_ders:
                        ders_kisitlari.at[tekrar_ders, diger_ders] = -500
                        ders_kisitlari.at[diger_ders, tekrar_ders] = -500

            for donem, donem_dersleri in donemler.items():
                for i in range(len(donem_dersleri)):
                    for j in range(i + 1, len(donem_dersleri)):
                        ders1 = donem_dersleri[i]
                        ders2 = donem_dersleri[j]
                        ders_kisitlari.at[ders1, ders2] = -1000  
                        ders_kisitlari.at[ders2, ders1] = -1000

        for bolum1, bolum2 in self.cap_bolumler:
    
            ders1_listesi = [
                ders for donem in self.bolumler.get(bolum1, {}).values()
                for ders in donem
            ]

            ders2_listesi = [
                ders for donem in self.bolumler.get(bolum2, {}).values()
                for ders in donem
            ]

            for ders1 in ders1_listesi:
                for ders2 in ders2_listesi:
                    if ders1 != ders2 and ders_kisitlari.at[ders1, ders2] == 0:
                        ders_kisitlari.at[ders1, ders2] = ceza_degeri
                        ders_kisitlari.at[ders2, ders1] = ceza_degeri

        return ders_kisitlari

    def takvim_hazirla(self):

        ders_kisitlari = self.ders_iliski_olustur()


        gunler = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"]
        zaman_dilimleri = ["08:30-11:30", "11:40-14:40", "14:50-17:50", "18:00-21:00"]
        sınıflar=self.siniflar
        bolumler=self.bolumler

        self.ders_bolum_sayilari = {}  
        ders_bolum_donem_map = defaultdict(lambda: {"departments": set(), "donemler": set()})

        for bolum, dönem_dersleri in self.bolumler.items():
            for donem, ders_listesi in dönem_dersleri.items():
                for ders in ders_listesi:
                    ders_bolum_donem_map[ders]["departments"].add(bolum)
                    ders_bolum_donem_map[ders]["donemler"].add(donem)

        for ders, data in ders_bolum_donem_map.items():
            bolum_sayisi = len(data["departments"])
            donemler = data["donemler"]

            if bolum_sayisi >= 2:
                ders_type = "multi"
                kapasite = 120
            elif any("Seçmeli" in d or "Tekrar" in d for d in donemler):
                ders_type = "elective"
                kapasite = 30
            else:
                ders_type = "core"
                kapasite = 60

            self.ders_bolum_sayilari[ders] = {
                "departments": list(data["departments"]),
                "type": ders_type,
                "bolum_sayisi": bolum_sayisi,
                "kapasite": kapasite
                }

        csp = CSPFinalTakvimi(ders_kisitlari, gunler, zaman_dilimleri,  sınıflar, self.ders_bolum_sayilari,bolumler)

        best_schedule = csp.solve()
        print("Best Final Schedule:")
        for exam in best_schedule:
            print(exam)

        if not best_schedule:
            print("Çözüm bulunamadı.")
            return

        root = tk.Tk()
        root.title("Final Takvimi")
        root.geometry("900x500")

        style = ttk.Style()
        style.configure("Treeview", rowheight=30)
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"))

        frame = ttk.Frame(root)
        frame.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")

        takvim_tablosu = ttk.Treeview(
            frame,
            columns=["Zaman"] + [str(gun) for gun in gunler],
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set
        )

        vsb.config(command=takvim_tablosu.yview)
        hsb.config(command=takvim_tablosu.xview)

        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        takvim_tablosu.pack(fill="both", expand=True)

        takvim_tablosu.heading("Zaman", text="Zaman")
        for gun in gunler:
            takvim_tablosu.heading(str(gun), text=f"Gün {gun}")

        for col in ["Zaman"] + [str(gun) for gun in gunler]:
            takvim_tablosu.column(col, width=350, anchor="center")

        satir_idleri = {}
        for zaman in zaman_dilimleri:
            values = [zaman] + [""] * len(gunler)
            satir_idleri[zaman] = takvim_tablosu.insert("", tk.END, values=values)

        for ders, (gun, zaman, sinif) in best_schedule.items():
            gun = str(gun)  
            if gun not in gunler:
                print(f"Uyarı: '{gun}' günü tanımsız, atlanıyor.")
                continue
            gun_index = gunler.index(gun) + 1  
            satir_id = satir_idleri[zaman]

            row_values = list(takvim_tablosu.item(satir_id, "values"))
            if row_values[gun_index]:
                row_values[gun_index] += f", {ders} ({sinif})"
            else:
                row_values[gun_index] = f"{ders} ({sinif})"

            takvim_tablosu.item(satir_id, values=row_values)
        takvim_tablosu.pack(fill=tk.BOTH, expand=True)


if __name__ == "__main__":
   root = tk.Tk()
   app = FinalTakvimUygulamasi(root)
   root.mainloop()
